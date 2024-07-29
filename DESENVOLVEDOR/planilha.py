from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from PSContBeam.DESENVOLVEDOR.erro import erro_percentual
def criar_resultados_xlsx(inf, sup, caminho, tensao):
    ROOT_FOLDER = Path(caminho).parent
    WORKBOOK    = ROOT_FOLDER / f"{tensao}.xlsx"

    workbook = Workbook()
    worksheet : Worksheet = workbook.active
    worksheet.title = tensao
    worksheet.cell(1,1, "Fibra Superior")
    worksheet.cell(1,2, "Fibra Inferior")
    for i in range(len(inf)):
        worksheet.cell(i+2, 1, float(sup[i]))
        worksheet.cell(i+2, 2, float(inf[i]))
    
    workbook.save(WORKBOOK)


def get_parms(parm, name, caminho):
    ROOT_FOLDER = Path(caminho).parent
    WORKBOOK    = ROOT_FOLDER / f"{name}.xlsx"

    workbook = Workbook()
    worksheet : Worksheet = workbook.active
    worksheet.cell(1,1, name)
    c = 2
    for i in parm:
        worksheet.cell(c, 1, float(i))
        c += 1
    
    workbook.save(WORKBOOK)