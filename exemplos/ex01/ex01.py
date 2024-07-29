from PSContBeam.Executor.leitura_arq import  leitura_dados
from PSContBeam.Tensao.Tensoes import Tensao
from openpyxl import  Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import  Cell
from PSContBeam.DESENVOLVEDOR.planilha import criar_resultados_xlsx
import numpy as np

# #  pegando valores das tensoes da planilha eletronica
excel : Workbook = load_workbook("G:\Meu Drive\TCC\Exemplos\ex01\ex.xlsx")
name_planilha = excel.sheetnames[0]
planilha : Worksheet = excel[name_planilha]

Planilha_result = {"ELS-F":{"inf": np.array([]), "sup": np.array([])},
                   "ELS-D":{"inf": np.array([]), "sup": np.array([])},
                   "ELU-ATO":{"inf": np.array([]), "sup": np.array([])}}

c = 2
for i in Planilha_result:
    for j in Planilha_result[i]:
        linha : tuple[Cell]
        for linha in planilha.iter_rows(min_row=44, max_row= 60, min_col= c, max_col= c):
            for cell in linha:
                Planilha_result[i][j] = np.append(Planilha_result[i][j], cell.value)
        c +=1
#----------------------------------------------------------------------------------------------------

dados = leitura_dados("retilineo.json")

tensao = Tensao(dados, size = 1, not_pp= True, part= 1)
#tensao.plot_momento_hiperestatico()
#tensao.plot_elu_ato()
#criar_resultados_xlsx(tensao.elu_ato().inf(), tensao.elu_ato().sup(), __file__, "ELU-ATO")
#print(tensao.els_f().sup())


# print(erro_max)
tensao.plot_els_f(sup= Planilha_result["ELS-F"]["sup"],
                  inf= Planilha_result["ELS-F"]["inf"])
tensao.plot_els_d(sup= Planilha_result["ELS-D"]["sup"],
                  inf= Planilha_result["ELS-D"]["inf"])
tensao.plot_elu_ato(sup= Planilha_result["ELU-ATO"]["sup"],
                    inf= Planilha_result["ELU-ATO"]["inf"])

